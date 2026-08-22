#!/usr/bin/env python3
"""
Generates all output files for neues-konzept from the full OCIANA corpus:
  - safaitic_invocations.xlsx   (already done by parse_safaitic.py, re-exported here)
  - safaitic_narrative.xlsx     (non-invocation Safaitic inscriptions with narrative)
  - safaitic_top100_interesting.txt
  - safaitic_top50_stories.txt
  - safaitic_full_corpus.xlsx   (all Safaitic, all lengths, for chapter I material)

Run from safaitic-research/:
  python3 archiv/neues-konzept/scripts/generate_outputs.py --xml ociana_corpus.xml
"""

import re
import sys
import argparse
from pathlib import Path
from collections import Counter

import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Deity detection (same as parse_safaitic.py) ──────────────────────────────

DEITIES = {
    "Allāt":        re.compile(r"\bh\s+(lt|ʾlt|ylt|yʾlt)\b"),
    "Rudā":         re.compile(r"\bh\s+(rḍw|rḍy|rḍ)\b"),
    "Baʿalshameen": re.compile(r"\bh\s+(bʿls¹mn|bʿls¹m)\b"),
    "Yaṯaʿ":        re.compile(r"\bh\s+(yṯʿ|ʾṯʿ)\b"),
    "Šaʿhaqam":     re.compile(r"\bh\s+(s²ʿhqm|s²ʿqm)\b"),
    "Dushara":      re.compile(r"\bh\s+(ds²r|ḏs²r)\b"),
    "Gaddaref":     re.compile(r"\bh\s+gdḍf\b"),
    "Gadʿawdh":     re.compile(r"\bh\s+gdʿwḏ\b"),
}

def has_invocation(translit):
    return any(p.search(translit) for p in DEITIES.values())

# ── Genealogy stripper ────────────────────────────────────────────────────────

_GENEAL = re.compile(
    r"^(?:By|For)\s+\S+(?:\s+(?:son|daughter|bint)\s+of\s+\S+)*"
    r"(?:\s+of\s+the\s+lineage\s+of\s+\S+)?[.,]?\s*",
    re.IGNORECASE,
)

def strip_genealogy(text):
    m = _GENEAL.match(text)
    return text[m.end():].strip() if m else text

# ── Scoring for "interesting" ranking ────────────────────────────────────────

_HIST = re.compile(r"\b(king|roman|nabat|war|battle|smote|slew|struggle|caesar|philippus|legion|soldier|warrior|persian)\b", re.I)
_ASTRO = re.compile(r"\b(libra|sagittarius|scorpio|moon|star|full moon|cosmical|solstice|zodiac)\b", re.I)
_EMO = re.compile(r"\b(grief|distraught|devastated|mourned|longed|wept|miserable|depressed|despair|broken heart|spirit was stripped)\b", re.I)
_SCENE = re.compile(r"\b(camped|pastured|watered|journey|migrated|inner desert|winter|summer|spring|year of|the year)\b", re.I)
_MULTI_DEITY = re.compile(r"\bO\s+\w", re.I)
_LACUNA = re.compile(r"----|\[\.\.\.\]|\[  \]")
_CURSE = re.compile(r"\b(blind|blindness|ejection from the grave|mutilation|madness|dumbness|lameness|curse|efface|scratch out|destroy)\b", re.I)

def score(core, deities_count=0):
    s = 0
    s += len(_HIST.findall(core)) * 5
    s += len(_ASTRO.findall(core)) * 4
    s += len(_EMO.findall(core)) * 3
    s += len(_SCENE.findall(core)) * 2
    s += deities_count * 3
    s += len(core) // 25
    s += len(_CURSE.findall(core)) * 2
    s -= len(_LACUNA.findall(core)) * 1
    return s

# ── Dedup ─────────────────────────────────────────────────────────────────────

def dedup_key(text):
    return re.sub(r"\s+", " ", re.sub(r"[^a-zA-Zʾʿ ]", "", text.lower()))[:100]

# ── XML loading ───────────────────────────────────────────────────────────────

def load_corpus(xml_path):
    with open(xml_path, "rb") as f:
        content = f.read()
    content = re.sub(b"<\\?xml[^?]*\\?>", b"", content)
    content = re.sub(b"<xs:schema>.*?</xs:schema>", b"", content, flags=re.DOTALL)
    content = re.sub(b"[\x00-\x08\x0b\x0c\x0e-\x1f]", b"", content)
    content = b"<corpus>" + content.strip() + b"</corpus>"
    return etree.fromstring(content)

# ── Excel helpers ─────────────────────────────────────────────────────────────

_HF   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BF   = Font(name="Calibri", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CTR  = Alignment(horizontal="center", vertical="center")
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def _write_sheet(ws, rows, columns, header_colour, col_widths):
    hfill = PatternFill("solid", fgColor=header_colour)
    for c, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = _HF; cell.fill = hfill
        cell.alignment = _CTR; cell.border = _THIN

    alt = PatternFill("solid", fgColor="EBF5FB")
    plain = PatternFill("solid", fgColor="FFFFFF")
    for r, rec in enumerate(rows, 2):
        fill = alt if r % 2 == 0 else plain
        for c, col in enumerate(columns, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(col, ""))
            cell.font = _BF; cell.fill = fill
            cell.alignment = _WRAP; cell.border = _THIN

    ws.freeze_panes = "A2"
    for c, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(col, 20)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 50

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xml", required=True, metavar="FILE")
    ap.add_argument("--out", default="archiv/neues-konzept/data")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    print(f"Loading {args.xml} …")
    root = load_corpus(Path(args.xml))
    all_ins = root.findall("inscription")
    safaitic = [i for i in all_ins if (i.findtext("script") or "").strip() == "Safaitic"]
    print(f"  Total: {len(all_ins):,}  |  Safaitic: {len(safaitic):,}")

    # ── Classify all Safaitic inscriptions ──────────────────────────────────

    invoc_records = []
    narr_records  = []
    full_records  = []   # every Safaitic inscription, for chapter-I material

    NAR_COLS = ["Inscription ID", "Location", "Core Narrative (EN)", "Full Translation", "Transliteration", "Score", "URL"]
    INV_COLS = ["Inscription ID", "Location", "Deities Invoked", "Full Translation", "Transliteration", "Score", "URL"]
    FULL_COLS = ["Inscription ID", "Location", "Core Length", "Type", "Core Narrative (EN)", "Full Translation", "URL"]

    seen_narr = set()

    for ins in safaitic:
        translit = (ins.findtext("transliteration") or "").strip()
        transl   = (ins.findtext("translation") or "").strip()
        siglum   = (ins.findtext("siglum") or "").strip()
        site     = (ins.findtext("site") or "").strip()
        region   = (ins.findtext("region") or "").strip()
        country  = (ins.findtext("country") or "").strip()
        location = ", ".join(p for p in [site, region, country] if p) or "—"
        url      = (ins.findtext("url") or "").strip()
        core     = strip_genealogy(transl)
        n_core   = len(core)

        # Classify type
        if has_invocation(translit):
            itype = "Invocation"
        elif n_core == 0:
            itype = "Signature"
        elif n_core < 30:
            itype = "Minimal"
        elif n_core < 80:
            itype = "Short narrative"
        else:
            itype = "Narrative"

        deities = [name for name, pat in DEITIES.items() if pat.search(translit)]
        sc = score(core, len(deities))

        full_records.append({
            "Inscription ID": siglum,
            "Location": location,
            "Core Length": n_core,
            "Type": itype,
            "Core Narrative (EN)": core[:400],
            "Full Translation": transl[:400],
            "URL": url,
        })

        if has_invocation(translit):
            invoc_records.append({
                "Inscription ID": siglum,
                "Location": location,
                "Deities Invoked": ", ".join(deities),
                "Full Translation": transl,
                "Transliteration": translit,
                "Score": sc,
                "URL": url,
            })
        elif n_core >= 30:
            dk = dedup_key(core)
            if dk not in seen_narr:
                seen_narr.add(dk)
                narr_records.append({
                    "Inscription ID": siglum,
                    "Location": location,
                    "Core Narrative (EN)": core,
                    "Full Translation": transl,
                    "Transliteration": translit,
                    "Score": sc,
                    "URL": url,
                })

    print(f"  Invocations: {len(invoc_records):,}")
    print(f"  Narratives (no invoc, core ≥ 30 chars, deduped): {len(narr_records):,}")
    print(f"  Full corpus rows: {len(full_records):,}")

    # ── Export narrative xlsx ────────────────────────────────────────────────
    narr_sorted = sorted(narr_records, key=lambda r: -r["Score"])
    narr_path = out / "safaitic_narrative.xlsx"
    df_n = pd.DataFrame(narr_sorted, columns=NAR_COLS)
    with pd.ExcelWriter(str(narr_path), engine="openpyxl") as writer:
        df_n.to_excel(writer, sheet_name="All Narratives", index=False)
    wb = load_workbook(str(narr_path))
    _write_sheet(wb["All Narratives"], narr_sorted, NAR_COLS, "1F497D",
                 {"Inscription ID": 16, "Location": 28, "Core Narrative (EN)": 60,
                  "Full Translation": 55, "Transliteration": 55, "Score": 8, "URL": 42})
    wb.save(str(narr_path))
    print(f"  → {narr_path}")

    # ── Export full corpus xlsx ──────────────────────────────────────────────
    full_path = out / "safaitic_full_corpus.xlsx"
    df_f = pd.DataFrame(full_records, columns=FULL_COLS)
    with pd.ExcelWriter(str(full_path), engine="openpyxl") as writer:
        df_f.to_excel(writer, sheet_name="All Safaitic", index=False)
        for itype in ["Signature", "Minimal", "Short narrative", "Narrative", "Invocation"]:
            sub = df_f[df_f["Type"] == itype]
            if not sub.empty:
                sub.to_excel(writer, sheet_name=itype[:31], index=False)
    wb2 = load_workbook(str(full_path))
    _write_sheet(wb2["All Safaitic"], full_records, FULL_COLS, "2C3E50",
                 {"Inscription ID": 14, "Location": 26, "Core Length": 10,
                  "Type": 16, "Core Narrative (EN)": 60, "Full Translation": 60, "URL": 42})
    wb2.save(str(full_path))
    print(f"  → {full_path}")

    # ── Top 50 longest narratives (txt) ──────────────────────────────────────
    top50 = sorted(narr_records, key=lambda r: -len(r["Core Narrative (EN)"]))[:50]
    top50_path = out / "safaitic_top50_stories.txt"
    with open(top50_path, "w", encoding="utf-8") as f:
        f.write("TOP 50 LÄNGSTE SAFAITISCHE NARRATIV-INSCHRIFTEN\n")
        f.write("Quelle: OCIANA-Corpus, University of Oxford\n")
        f.write("=" * 70 + "\n\n")
        for i, rec in enumerate(top50, 1):
            f.write(f"#{i}  {rec['Inscription ID']}  ({len(rec['Core Narrative (EN)'])} Zeichen Kern)\n")
            f.write(f"URL: {rec['URL']}\n\n")
            f.write(f"Vollständige Übersetzung:\n{rec['Full Translation']}\n\n")
            f.write(f"Kern (ohne Genealogie):\n{rec['Core Narrative (EN)']}\n\n")
            f.write("-" * 70 + "\n\n")
    print(f"  → {top50_path}")

    # ── Top 100 most interesting (txt) ────────────────────────────────────────
    top100 = sorted(narr_records + invoc_records, key=lambda r: -r["Score"])
    seen100 = set(); top100_dedup = []
    for rec in top100:
        dk = dedup_key(rec.get("Core Narrative (EN)", rec.get("Full Translation", "")))
        if dk not in seen100:
            seen100.add(dk); top100_dedup.append(rec)
        if len(top100_dedup) == 100:
            break

    top100_path = out / "safaitic_top100_interesting.txt"
    with open(top100_path, "w", encoding="utf-8") as f:
        f.write("TOP 100 INTERESSANTESTE SAFAITISCHE INSCHRIFTEN\n")
        f.write("Quelle: OCIANA-Corpus, University of Oxford\n")
        f.write("=" * 70 + "\n\n")
        for i, rec in enumerate(top100_dedup, 1):
            core = rec.get("Core Narrative (EN)", strip_genealogy(rec.get("Full Translation","")))
            f.write(f"#{i}  {rec['Inscription ID']}  (Score: {rec['Score']}, {len(core)} Zeichen)\n")
            f.write(f"URL: {rec['URL']}\n\n")
            f.write(f"Kern:\n{core}\n\n")
            f.write("-" * 70 + "\n\n")
    print(f"  → {top100_path}")

    print("\nAlle Ausgabedateien generiert.")

if __name__ == "__main__":
    main()
