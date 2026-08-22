#!/usr/bin/env python3
"""
build_xlsx.py — erzeugt die Auswahltabelle aus dem Scoring-Ergebnis.

Pipeline:
  1) scripts/score.py        -> scored.pkl   (erwartet safaitic_narrative.xlsx im CWD)
  2) scripts/build_xlsx.py   -> data/safaitic_gedichtband_auswahl.xlsx

Blaetter: Lies mich | Shortlist (kuratiert) | Gewichtung (editierbar) | Longlist (Live-Formel).
Vom Repo-Wurzelverzeichnis aus aufrufen: python3 scripts/build_xlsx.py
"""
import os
import pandas as pd, numpy as np, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_pickle('scored.pkl').copy()

# ---- de-cluster near-identical formulae (signature = core minus names/special tokens)
def sig(s):
    toks = re.findall(r"[a-zA-Z]+", str(s).lower())
    stop_names = {'by','son','of'}
    # drop tokens that are clearly name-fragments (very short caps handled by lower) -> keep content words
    keep = [w for w in toks if w not in stop_names and len(w) > 2]
    # remove obvious proper-name leftovers: keep only words appearing in a content vocab proxy -> just collapse
    return ' '.join(keep)[:90]
df['sig'] = df['core'].apply(sig)
df['rk_in_sig'] = df.groupby('sig').cumcount()

# arc mapping from primary theme
arc_map = {
 'Dry Season & Camping': ('Dürre & Lager', 1),
 'Journey & Travel': ('Wege & Wasser', 2),
 'Watching & Scouting': ('Spähen & Warten', 3),
 'Pastoral & Herding': ('Weide & Herde', 4),
 'Raiding & Warfare': ('Raub & Krieg', 5),
 'Disease & Injury': ('Krankheit & Versehrung', 6),
 'Death & Murder': ('Tod & Gewalt', 7),
 'Grief & Mourning': ('Klage', 8),
 'Fear & Anxiety': ('Angst', 9),
 'Longing & Homesickness': ('Sehnsucht & Heimweh', 10),
 'Unusual & Unique': ('Zeichen & Seltsames', 11),
 'Historical Event': ('Verankerung (Jahr & König)', 12),
}
prim = df['Theme'].astype(str).str.split(';').str[0].str.strip()
df['Arc'] = prim.map(lambda x: arc_map.get(x, ('Sonstiges', 13))[0])
df['ArcOrder'] = prim.map(lambda x: arc_map.get(x, ('Sonstiges', 13))[1])
df['Refrain'] = (df['meta'] > 0).astype(int)  # Fluch/Schrift = wiederkehrender Refrain

# ---- SHORTLIST: max 2 per formula-signature, protect rare arcs, target ~120
short = df[df['rk_in_sig'] < 2].copy()
# guarantee inclusion of all small arcs
parts = []
for arc, sub in short.groupby('Arc'):
    n = len(sub)
    take = min(n, max(6, int(round(n*0.10))))   # at least 6 per arc, ~10% of big arcs
    parts.append(sub.sort_values('LiteraryScore', ascending=False).head(take))
short = pd.concat(parts).sort_values(['ArcOrder','LiteraryScore'], ascending=[True,False]).reset_index(drop=True)

print('Shortlist:', len(short), '| Arc-Verteilung:')
print(short['Arc'].value_counts().reindex([arc_map[k][0] for k in arc_map]).fillna(0).astype(int))

# ================= WRITE WORKBOOK =================
wb = Workbook(); wb.remove(wb.active)
F='Arial'
hdr_fill = PatternFill('solid', fgColor='1F3864'); hdr_font=Font(name=F,bold=True,color='FFFFFF',size=10)
sub_fill = PatternFill('solid', fgColor='D6DCE4')
thin = Side(style='thin', color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top'); top=Alignment(vertical='top')

def style_header(ws, ncol, row=1):
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c); cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal='left',vertical='center'); cell.border=border

# ---- Sheet 1: Lies mich
ws0 = wb.create_sheet('Lies mich')
intro = [
 ('Safaitic — Gedichtband: Arbeitsmaterial', 14, True),
 ('', 10, False),
 ('Zweck: literarische, verknappte, klanghafte Auswahl; verankert, nicht wissenschaftlich.', 10, False),
 ('Die Auswahl ist kuratiert und sagt damit auch über uns etwas aus, nicht nur über die Steppe.', 10, False),
 ('', 10, False),
 ('Blätter:', 11, True),
 ('• Shortlist (kuratiert) — entkernte, jahresbogen-sortierte Auswahl als Bandgerüst.', 10, False),
 ('• Longlist (gerankt) — alle 2.451 Inschriften mit Score-Bausteinen.', 10, False),
 ('• Gewichtung — hier Gewichte ändern; LiteraryScore rechnet in der Longlist neu.', 10, False),
 ('', 10, False),
 ('Score-Bausteine:', 11, True),
 ('emo=Affekt, scene=konkretes Bild, rarity=Seltenheit des Themas, meta=Fluch/Schrift-Selbstbezug,', 10, False),
 ('anchor=Jahr/König (Verankerung), brevity=verknappte Dichte, integrity=wenig Lücken (----),', 10, False),
 ('clarity=wenig unsichere Lesung ({ }/[ ]), geneal_pen=Abzug für lange Ahnenketten.', 10, False),
 ('', 10, False),
 ('Refrain=1: enthält Fluch-/Segensformel — als wiederkehrendes Motiv im Band nutzbar.', 10, False),
 ('Hinweis: Übersetzungen aus OCIANA (englisch). Deutsche Nachdichtung ist ein weiterer Schritt.', 10, False),
]
for i,(txt,sz,b) in enumerate(intro,1):
    c=ws0.cell(row=i,column=1,value=txt); c.font=Font(name=F,size=sz,bold=b)
ws0.column_dimensions['A'].width=110

# ---- Sheet 2: Shortlist
ws=wb.create_sheet('Shortlist (kuratiert)')
cols=['ArcOrder','Arc','Refrain','LiteraryScore','Theme','Translation','Location','Inscription ID','URL']
labels=['#','Phase (Jahresbogen)','Refrain','Score','Thema','Übersetzung (OCIANA)','Fundort','ID','URL']
ws.append(labels); style_header(ws,len(labels))
for _,r in short.iterrows():
    ws.append([r['ArcOrder'],r['Arc'],r['Refrain'],r['LiteraryScore'],r['Theme'],
               r['Translation'],r['Location'],r['Inscription ID'],r['URL']])
widths=[5,22,9,8,26,70,28,18,30]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
    for c in row: c.font=Font(name=F,size=10); c.alignment=wrap; c.border=border
    if row[1].value and short[short['Arc']==row[1].value] is not None: pass
ws.freeze_panes='A2'

# ---- Sheet 3: Gewichtung (editable weights)
wsg=wb.create_sheet('Gewichtung')
wsg.append(['Baustein','Gewicht']); style_header(wsg,2)
weights=[('emo',3.0),('scene',2.5),('rarity',3.0),('meta',1.5),('anchor',1.0),
         ('brevity',2.5),('integrity',2.0),('clarity',1.5),('geneal_pen',-2.0)]
for name,w in weights:
    wsg.append([name,w])
for row in wsg.iter_rows(min_row=2,max_row=wsg.max_row):
    row[0].font=Font(name=F,size=10); row[1].font=Font(name=F,size=10,color='0000FF')
    row[1].fill=PatternFill('solid',fgColor='FFFF00')
    for c in row: c.border=border
wsg.column_dimensions['A'].width=14; wsg.column_dimensions['B'].width=12
wmap={name:f'Gewichtung!$B${i+2}' for i,(name,_) in enumerate(weights)}

# ---- Sheet 4: Longlist with LIVE formula score
wl=wb.create_sheet('Longlist (gerankt)')
comp=['emo','scene','rarity','meta','anchor','brevity','integrity','clarity','geneal_pen']
head=['LiteraryScore']+comp+['Theme','Arc','Refrain','core_words','lacuna','Translation','Location','Inscription ID','URL']
wl.append(head); style_header(wl,len(head))
dl=df.sort_values('LiteraryScore',ascending=False).reset_index(drop=True)
# column letters for components
ccol={name:get_column_letter(2+i) for i,name in enumerate(comp)}  # emo=B...
for ridx,(_,r) in enumerate(dl.iterrows(), start=2):
    formula = '=' + '+'.join(f'{ccol[n]}{ridx}*{wmap[n]}' for n in comp)
    wl.append([formula]+[r[n] for n in comp]+[r['Theme'],r['Arc'],r['Refrain'],
               r['core_words'],r['lacuna'],r['Translation'],r['Location'],r['Inscription ID'],r['URL']])
lw=[12]+[7]*9+[24,20,8,9,7,66,26,16,28]
for i,w in enumerate(lw,1): wl.column_dimensions[get_column_letter(i)].width=w
for row in wl.iter_rows(min_row=2,max_row=wl.max_row):
    for c in row: c.font=Font(name=F,size=9); c.alignment=top; c.border=border
    row[16].alignment=wrap
wl.freeze_panes='A2'

os.makedirs('data', exist_ok=True)
wb.save('data/safaitic_gedichtband_auswahl.xlsx')
print('gespeichert.')
