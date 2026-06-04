#!/usr/bin/env python3
"""
OCIANA Safaitic Invocation Extractor
=====================================
Parses the OCIANA bulk XML corpus, extracts all Safaitic inscriptions
containing divine invocations (vocative h + deity name), and exports a
styled .xlsx file with two classification columns:
  - Request Type  : the specific English meaning of the request word
  - Category      : the broader scholarly grouping

Usage:
    python3 parse_safaitic.py --xml ociana_corpus.xml
    python3 parse_safaitic.py --sample          # bundled test data
"""

import re
import sys
import argparse
from pathlib import Path

import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Deity detection ──────────────────────────────────────────────────────────
# In the OCIANA corpus the vocative particle 'h' is a standalone word followed
# by the deity name (e.g. 'h lt' = "O Allāt"), not prefixed with a dash.

DEITIES = {
    "Allāt":        re.compile(r"\bh\s+(lt|ʾlt|ylt|yʾlt)\b"),
    "Rudā":         re.compile(r"\bh\s+(rḍw|rḍy|rḍ)\b"),
    "Baʿalshameen": re.compile(r"\bh\s+(bʿls¹mn|bʿls¹m)\b"),
    "Yaṯaʿ":        re.compile(r"\bh\s+(yṯʿ|ʾṯʿ)\b"),
    "Šaʿhaqam":     re.compile(r"\bh\s+(s²ʿhqm|s²ʿqm)\b"),
    "Dushara":      re.compile(r"\bh\s+(ds²r|ḏs²r)\b"),
    "Gaddaref":     re.compile(r"\bh\s+gdḍf\b"),
    "Gadʿawdh":     re.compile(r"\bh\s+gdʿwḏ\b"),
}

# Matches any deity invocation and captures up to 60 chars of what follows,
# so _request_info can scan past conjunction/particle words to find the request.
_ANY_DEITY = re.compile(
    r"\bh\s+(?:lt|ʾlt|ylt|yʾlt"
    r"|rḍw|rḍy|rḍ"
    r"|bʿls¹mn|bʿls¹m"
    r"|yṯʿ|ʾṯʿ"
    r"|s²ʿhqm|s²ʿqm"
    r"|ds²r|ḏs²r"
    r"|gdḍf|gdʿwḏ)"
    r"\s+(.{1,80}?)(?=\b(?:w|f)\s+h\s+|\Z)"  # up to next invocation or end
)

# Particles / conjunctions / deity names to skip when scanning for the request
# word.  Deity names appear here because paired invocations like "h lt w ds²r"
# would otherwise return "ds²r" as the request.
_PARTICLES = {
    "w", "f", "l", "l-", "w-", "f-", "b", "b-", "mn", "ʿl-", "ʿl",
    "ʾl-", "ʾl", "ḏ-", "ḏ", "m-", "kl", "mr",
    # Deity name forms
    "lt", "ʾlt", "ylt", "yʾlt",
    "rḍw", "rḍy", "rḍ",
    "bʿls¹mn", "bʿls¹m",
    "yṯʿ", "ʾṯʿ",
    "s²ʿhqm", "s²ʿqm",
    "ds²r", "ḏs²r",
    "gdḍf", "gdʿwḏ",
}

# ── Request-word → English label ─────────────────────────────────────────────
# Maps the first Safaitic word after the deity name to a human-readable label.
REQUEST_LABELS = {
    # Security & Protection
    "s¹lm":    "Security",
    "s¹[l]m":  "Security",
    "s¹{l}m":  "Security",
    "s¹]lm":   "Security",
    "s¹(l)m":  "Security",
    "wqyt":    "Preservation",
    "flṭ":     "Deliverance",
    "ġwṯ":     "Succour / Help",
    "ḥnn":     "Mercy / Grace",
    "brʾ":     "Healing",
    "ḥmyt":    "Protection",
    # Vengeance & Cursing
    "ṯʾr":     "Vengeance",
    "nqm":     "Vengeance",
    "nqmt":    "Retribution",
    "ʿwr":     "Blindness on enemy",
    "nqʾt":    "Evil eye on enemy",
    "lʿn":     "Curse",
    "ʿyr":     "Shame on enemy",
    "ʿyrt":    "Shame on enemy",
    "fṣyt":    "Divine judgment",
    "ʿqbt":    "Punishment",
    # Prosperity & Gain
    "ġnmt":    "Plunder / Spoils",
    "ġnyt":    "Wealth",
    "s¹ʿd":    "Happiness / Fortune",
    "qbll":    "Benevolence",
    "hb":      "Grant [gift]",
    "whb":     "Grant [gift]",
    "ʾws¹":    "Gift",
    "mgdt":    "Gift",
    "ġnm":     "Flock / Gain",
    "ġyr":     "Zealous protection",
    # Relief & Wellbeing
    "rwḥ":     "Relief / Ease",
    "r{w}ḥ":   "Relief / Ease",
    "r(w)ḥ":   "Relief / Ease",
    "rw":      "Relief / Ease",          # damaged rwḥ
    "ġyrt":    "Zealous wellbeing",
    "mḥlt":    "Respite",
    # Damaged / variant spellings of security
    "s¹":      "Security",
    "s¹l":     "Security",
    "s¹l(m":   "Security",
    "s¹)lm":   "Security",
    "slm":     "Security",
    # Damaged / variant spellings of prosperity
    "s¹ʿ":     "Happiness / Fortune",
    "s¹ʿ(d":   "Happiness / Fortune",
    "ġny":     "Wealth",
    "ġrt":     "Zealous protection",
    "qbl":     "Benevolence",
    # Verbal / variant forms of vengeance/cursing
    "yʿwr":    "Blindness on enemy",
    "wr":      "Blindness on enemy",
    "ʿw":      "Blindness on enemy",
    "fṣy":     "Divine judgment",
    "nqʾ":     "Evil eye on enemy",
    "qʾt":     "Evil eye on enemy",
    "qmt":     "Retribution",
    "qm":      "Vengeance",
}

# ── Request-word → Category ──────────────────────────────────────────────────
REQUEST_CATEGORIES = {
    # Security & Protection
    "s¹lm":    "Security & Protection",
    "s¹[l]m":  "Security & Protection",
    "s¹{l}m":  "Security & Protection",
    "s¹]lm":   "Security & Protection",
    "s¹(l)m":  "Security & Protection",
    "wqyt":    "Security & Protection",
    "flṭ":     "Security & Protection",
    "ġwṯ":     "Security & Protection",
    "ḥnn":     "Security & Protection",
    "brʾ":     "Security & Protection",
    "ḥmyt":    "Security & Protection",
    # Vengeance & Cursing
    "ṯʾr":     "Vengeance & Cursing",
    "nqm":     "Vengeance & Cursing",
    "nqmt":    "Vengeance & Cursing",
    "ʿwr":     "Vengeance & Cursing",
    "nqʾt":    "Vengeance & Cursing",
    "lʿn":     "Vengeance & Cursing",
    "ʿyr":     "Vengeance & Cursing",
    "ʿyrt":    "Vengeance & Cursing",
    "fṣyt":    "Vengeance & Cursing",
    "ʿqbt":    "Vengeance & Cursing",
    # Prosperity & Gain
    "ġnmt":    "Prosperity & Gain",
    "ġnyt":    "Prosperity & Gain",
    "s¹ʿd":    "Prosperity & Gain",
    "qbll":    "Prosperity & Gain",
    "hb":      "Prosperity & Gain",
    "whb":     "Prosperity & Gain",
    "ʾws¹":    "Prosperity & Gain",
    "mgdt":    "Prosperity & Gain",
    "ġnm":     "Prosperity & Gain",
    "ġyr":     "Prosperity & Gain",
    # Relief & Wellbeing
    "rwḥ":     "Relief & Wellbeing",
    "r{w}ḥ":   "Relief & Wellbeing",
    "r(w)ḥ":   "Relief & Wellbeing",
    "rw":      "Relief & Wellbeing",
    "ġyrt":    "Relief & Wellbeing",
    "mḥlt":    "Relief & Wellbeing",
    # Damaged security variants
    "s¹":      "Security & Protection",
    "s¹l":     "Security & Protection",
    "s¹l(m":   "Security & Protection",
    "s¹)lm":   "Security & Protection",
    "slm":     "Security & Protection",
    # Damaged prosperity variants
    "s¹ʿ":     "Prosperity & Gain",
    "s¹ʿ(d":   "Prosperity & Gain",
    "ġny":     "Prosperity & Gain",
    "ġrt":     "Prosperity & Gain",
    "qbl":     "Prosperity & Gain",
    # Verbal/damaged vengeance forms
    "yʿwr":    "Vengeance & Cursing",
    "wr":      "Vengeance & Cursing",
    "ʿw":      "Vengeance & Cursing",
    "fṣy":     "Vengeance & Cursing",
    "nqʾ":     "Vengeance & Cursing",
    "qʾt":     "Vengeance & Cursing",
    "qmt":     "Vengeance & Cursing",
    "qm":      "Vengeance & Cursing",
}

_CAT_ORDER = [
    "Security & Protection",
    "Vengeance & Cursing",
    "Prosperity & Gain",
    "Relief & Wellbeing",
    "Mixed",
    "Other",
]

_CAT_COLOURS = {
    "Security & Protection": "1F497D",   # dark blue
    "Vengeance & Cursing":   "C0392B",   # deep red
    "Prosperity & Gain":     "27AE60",   # green
    "Relief & Wellbeing":    "E67E22",   # orange
    "Mixed":                 "8E44AD",   # purple
    "Other":                 "7F8C8D",   # grey
}

# ── Columns ──────────────────────────────────────────────────────────────────
COLUMNS = [
    "Inscription ID",
    "Location",
    "Deities Invoked",
    "Request Type",
    "Category",
    "Invocation (English)",
    "Full Translation",
    "Safaitic Transliteration",
    "URL",
]

COL_WIDTHS = {
    "Inscription ID":        16,
    "Location":              28,
    "Deities Invoked":       24,
    "Request Type":          28,
    "Category":              26,
    "Invocation (English)":  55,
    "Full Translation":      55,
    "Safaitic Transliteration": 55,
    "URL":                   45,
}

# ── Parsing ──────────────────────────────────────────────────────────────────

_INVOC_SPLIT = re.compile(r"(?:(?:[,.]\s*)?(?:So\s+)?)(\bO\b.+)", re.IGNORECASE | re.DOTALL)


def _extract_invocation(translation: str) -> str:
    """Return from the first 'O [Deity]' onward, capitalised."""
    m = _INVOC_SPLIT.search(translation)
    if m:
        rest = m.group(1).strip()
        return rest[0].upper() + rest[1:] if rest else rest
    # Fallback: 'and may / So may'
    m2 = re.search(r"(?:[,.]\s*)?(?:So\s+)?(?:and\s+)?may\s+.+", translation, re.IGNORECASE | re.DOTALL)
    if m2:
        rest = m2.group(0).strip().lstrip(",. ")
        return rest[0].upper() + rest[1:] if rest else rest
    return translation


def _first_request_word(context: str) -> str:
    """Return the first non-particle word from the post-deity context string."""
    for tok in re.split(r"[\s{}\[\]<>.,;]+", context):
        tok = tok.strip(".,;-{}[]()")
        if tok and tok.lower() not in _PARTICLES and len(tok) > 1:
            return tok
    return ""


def _request_info(transliteration: str) -> tuple[str, str]:
    """Return (request_type_label, category) derived from the request word(s)."""
    req_words = []
    for m in _ANY_DEITY.finditer(transliteration):
        context = m.group(1) if m.lastindex else ""
        w = _first_request_word(context)
        if w:
            req_words.append(w)

    if not req_words:
        return ("—", "Other")

    labels = []
    cats = set()
    for w in req_words:
        lbl = REQUEST_LABELS.get(w)
        cat = REQUEST_CATEGORIES.get(w)
        if lbl:
            labels.append(lbl)
        if cat:
            cats.add(cat)

    label_str = "; ".join(dict.fromkeys(labels)) if labels else req_words[0]
    if len(cats) == 0:
        category = "Other"
    elif len(cats) == 1:
        category = next(iter(cats))
    else:
        category = "Mixed"

    return (label_str, category)


def _load_corpus(xml_path: Path) -> etree._Element:
    with open(xml_path, "rb") as f:
        content = f.read()
    content = re.sub(b"<\\?xml[^?]*\\?>", b"", content)
    content = re.sub(b"<xs:schema>.*?</xs:schema>", b"", content, flags=re.DOTALL)
    content = re.sub(b"[\x00-\x08\x0b\x0c\x0e-\x1f]", b"", content)
    content = b"<corpus>" + content.strip() + b"</corpus>"
    return etree.fromstring(content)


def parse_corpus(xml_path: Path) -> list[dict]:
    print(f"  Parsing {xml_path} …")
    root = _load_corpus(xml_path)
    inscriptions = root.findall("inscription")
    print(f"  Total inscriptions in file: {len(inscriptions):,}")

    records = []
    for ins in inscriptions:
        script = (ins.findtext("script") or "").strip()
        if script != "Safaitic":
            continue

        translit = (ins.findtext("transliteration") or "").strip()
        transl   = (ins.findtext("translation") or "").strip()

        deities_found = [name for name, pat in DEITIES.items() if pat.search(translit)]
        if not deities_found:
            continue

        siglum   = (ins.findtext("siglum") or "").strip()
        site     = (ins.findtext("site") or "").strip()
        region   = (ins.findtext("region") or "").strip()
        country  = (ins.findtext("country") or "").strip()
        location = ", ".join(p for p in [site, region, country] if p) or "—"
        url      = (ins.findtext("url") or "").strip()

        request_type, category = _request_info(translit)
        invocation = _extract_invocation(transl)

        records.append({
            "Inscription ID":        siglum,
            "Location":              location,
            "Deities Invoked":       ", ".join(deities_found),
            "Request Type":          request_type,
            "Category":              category,
            "Invocation (English)":  invocation,
            "Full Translation":      transl,
            "Safaitic Transliteration": translit,
            "URL":                   url,
        })

    return records


def parse_sample(sample_dir: Path) -> list[dict]:
    """Parse the bundled EpiDoc TEI sample files (legacy format)."""
    TEI_NS = "http://www.tei-c.org/ns/1.0"
    NS = {"tei": TEI_NS}

    def _text(el):
        return " ".join(t.strip() for t in el.itertext() if t.strip()) if el is not None else ""

    records = []
    for xml_path in sorted(sample_dir.glob("*.xml")):
        try:
            tree = etree.parse(str(xml_path))
        except etree.XMLSyntaxError:
            continue
        root = tree.getroot()
        idno = root.find(".//tei:idno[@type='filename']", NS)
        siglum = idno.text.strip() if idno is not None and idno.text else xml_path.stem
        settlement = root.find(".//tei:settlement", NS)
        location = (settlement.text or "").strip() if settlement is not None else "—"
        edition = root.find(".//tei:div[@type='edition']", NS)
        translit = _text(edition)
        transl_div = root.find(".//tei:div[@type='translation']", NS)
        transl = _text(transl_div)

        # Use translation-based detection for sample data
        deity_trans = re.compile(
            r"\bAllāt\b|\bRudā\b|\bShams\b", re.IGNORECASE
        )
        if not deity_trans.search(transl):
            continue

        deities = [d for d in ["Allāt", "Rudā", "Shams"] if re.search(rf"\b{d}\b", transl, re.IGNORECASE)]

        # Simple category from translation
        cat_map = [
            ("Vengeance & Cursing",   re.compile(r"\b(veng|curse|retrib|evil|enemy|punish)\b", re.I)),
            ("Security & Protection", re.compile(r"\b(protect|peace|security|help|save|guard)\b", re.I)),
            ("Prosperity & Gain",     re.compile(r"\b(pasture|rain|gain|fortune|plunder|wealth)\b", re.I)),
            ("Relief & Wellbeing",    re.compile(r"\b(relief|ease|wellbeing|benevolence)\b", re.I)),
        ]
        cats = [cat for cat, pat in cat_map if pat.search(transl)]
        category = cats[0] if len(cats) == 1 else ("Mixed" if len(cats) > 1 else "Other")
        request_type = "; ".join(cats) if cats else "—"

        invocation = _extract_invocation(transl)
        records.append({
            "Inscription ID":           siglum,
            "Location":                 location,
            "Deities Invoked":          ", ".join(deities),
            "Request Type":             request_type,
            "Category":                 category,
            "Invocation (English)":     invocation,
            "Full Translation":         transl,
            "Safaitic Transliteration": translit,
            "URL":                      "",
        })
    return records


# ── Excel export ─────────────────────────────────────────────────────────────

_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT    = Font(name="Calibri", size=10)
_WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN         = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _lighten(hex_colour: str, factor: float = 0.85) -> str:
    r = int(hex_colour[0:2], 16); g = int(hex_colour[2:4], 16); b = int(hex_colour[4:6], 16)
    return "{:02X}{:02X}{:02X}".format(
        int(r + (255 - r) * factor),
        int(g + (255 - g) * factor),
        int(b + (255 - b) * factor),
    )


def _header_row(ws, colour: str):
    fill = PatternFill("solid", fgColor=colour)
    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = _HEADER_FONT; cell.fill = fill
        cell.alignment = _CENTER_ALIGN; cell.border = _THIN


def _data_rows(ws, rows: list[dict], alt_colour: str):
    alt   = PatternFill("solid", fgColor=alt_colour)
    plain = PatternFill("solid", fgColor="FFFFFF")
    for r, rec in enumerate(rows, 2):
        fill = alt if r % 2 == 0 else plain
        for c, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(col, ""))
            cell.font = _BODY_FONT; cell.fill = fill
            cell.alignment = _WRAP_ALIGN; cell.border = _THIN


def _freeze_size(ws):
    ws.freeze_panes = "A2"
    for c, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[col]
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 55


def export_excel(records: list[dict], output_path: Path):
    if not records:
        print("  No matching inscriptions — nothing to export."); return

    df = pd.DataFrame(records, columns=COLUMNS)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # Summary sheet
        summary_rows = []
        for cat in _CAT_ORDER:
            sub = df[df["Category"] == cat]
            if sub.empty:
                continue
            top_d = (
                sub["Deities Invoked"].str.split(", ").explode()
                .value_counts().head(3).index.tolist()
            )
            top_r = (
                sub["Request Type"].str.split("; ").explode()
                .value_counts().head(3).index.tolist()
            )
            summary_rows.append({
                "Category":        cat,
                "Count":           len(sub),
                "% of Total":      f"{100*len(sub)/len(df):.1f}%",
                "Top Deities":     ", ".join(top_d),
                "Top Requests":    ", ".join(top_r),
            })
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="All Invocations", index=False)
        for cat in _CAT_ORDER:
            sub = df[df["Category"] == cat]
            if not sub.empty:
                sub.to_excel(writer, sheet_name=cat[:31], index=False)

    wb = load_workbook(str(output_path))

    # Format Summary
    ws = wb["Summary"]
    ws.freeze_panes = "A2"
    dark = PatternFill("solid", fgColor="2C3E50")
    for cell in ws[1]:
        cell.font = _HEADER_FONT; cell.fill = dark
        cell.alignment = _CENTER_ALIGN; cell.border = _THIN
    for r in range(2, ws.max_row + 1):
        cat_val = ws.cell(r, 1).value or ""
        cat_fill = PatternFill("solid", fgColor=_CAT_COLOURS.get(cat_val, "7F8C8D"))
        row_fill = PatternFill("solid", fgColor="ECF0F1" if r % 2 == 0 else "FFFFFF")
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = cat_fill if c == 1 else row_fill
            cell.font = Font(name="Calibri", size=11,
                             bold=(c == 1), color="FFFFFF" if c == 1 else "2C3E50")
            cell.alignment = _CENTER_ALIGN; cell.border = _THIN
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 32

    # Format All Invocations
    ws_all = wb["All Invocations"]
    _header_row(ws_all, "2C3E50")
    _data_rows(ws_all, records, "EBF5FB")
    _freeze_size(ws_all)

    # Format per-category sheets
    for cat in _CAT_ORDER:
        name = cat[:31]
        if name not in wb.sheetnames:
            continue
        ws_cat = wb[name]
        colour = _CAT_COLOURS.get(cat, "2C3E50")
        subset = [r for r in records if r["Category"] == cat]
        _header_row(ws_cat, colour)
        _data_rows(ws_cat, subset, _lighten(colour))
        _freeze_size(ws_cat)

    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)
    wb.save(str(output_path))
    print(f"\n  Exported {len(records):,} inscriptions → {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="OCIANA Safaitic Invocation Extractor")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--xml",    metavar="FILE", help="OCIANA bulk corpus XML file")
    src.add_argument("--sample", action="store_true", help="Use bundled sample EpiDoc XML files")
    ap.add_argument("--output",  metavar="FILE", default="safaitic_invocations.xlsx")
    args = ap.parse_args()

    print("=" * 60)
    print("  OCIANA Safaitic Invocation Extractor")
    print("=" * 60)

    if args.xml:
        xml_path = Path(args.xml)
        if not xml_path.exists():
            print(f"ERROR: file not found: {xml_path}"); sys.exit(1)
        print(f"\n[1/3] Loading corpus from {xml_path}")
        records = parse_corpus(xml_path)
    else:
        sample_dir = Path(__file__).parent / "sample_data"
        print(f"\n[1/3] Loading sample data from {sample_dir}")
        records = parse_sample(sample_dir)

    print(f"\n[2/3] Parsing complete — {len(records):,} invocation inscriptions found")
    if records:
        cat_counts = {}
        for r in records:
            cat_counts[r["Category"]] = cat_counts.get(r["Category"], 0) + 1
        print("\n  Category breakdown:")
        for cat in _CAT_ORDER:
            if cat in cat_counts:
                print(f"    {cat:<30} {cat_counts[cat]:>6,}")

    print(f"\n[3/3] Exporting to {args.output} …")
    export_excel(records, Path(args.output))
    print("\nDone.")


if __name__ == "__main__":
    main()
